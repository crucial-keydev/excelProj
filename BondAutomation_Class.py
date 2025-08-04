from openpyxl import workbook, load_workbook, worksheet
from functs import get_amount, get_perc,roundup, getpayment, getWords, turntosentence, fitintobox, rating_prf,rating_sty, get_bond_type
import datetime
import json

with open('config.json', 'r') as f:
    config = json.load(f)

ctc = config["ctc"]
ctc_date = config["ctc_date"]
manager = config["manager"]
ctc_manager = config["ctc_manager"]
ctc_manager_date = config["ctc_manager_date"]


today = datetime.datetime.now()
year = int(today.year)

class Bond:

    con_prop = {} 
        
    def __init__(self):
        
        self.get_contractor()
        self.get_prop_addr()
        self.get_agency()
        self.get_date()
        self.get_contract_name()
        self.get_contract_amount()
        self.get_percent()
        self.get_coverage()
        self.get_coverage_in_words()
        self.get_bond_no()
        self.get_bond_type()
        self.get_or_no()

    def get_coverage(self):
        self.coverage = int(roundup(self.contract_amount * self.perc))
    
    def get_contractor(self):
        self.contractor = input('Contractor: ')

    def get_prop_addr(self):
        if self.contractor in Bond.con_prop:
            prop, prop_address = Bond.con_prop[self.contractor]
            
        else:
            prop = input("Proprietor: ")
            prop_address = input("Address: ")
            Bond.con_prop[self.contractor] = [prop, prop_address]
        self.prop = prop
        self.prop_address = prop_address

    def get_contract_name(self):
        self.contract_name = input('Contract Name: ')

    def get_date(self):
        self.date = input("Date: ")

    def get_contract_amount(self):
        self.contract_amount = get_amount()
    
    def get_percent(self):
        self.perc = get_perc()/100

    def get_agency(self):
        self.agency = input('Agency: ')

    def get_coverage_in_words(self):
        self.coverage_in_words = getWords(self.coverage)

    def get_bond_no(self):
        self.bond_no = input('Bond number: ')

    def get_bond_type(self):
        bondtype = input('STY or PRF or WARR or MOB: ').upper()
        while True:
            try:
                bondtype == 'STY' or bondtype == 'PRF' or bondtype == 'WARR' or bondtype =='MOB' 
            except:
                print('Invalid type')

            if bondtype == 'STY' or bondtype == 'PRF' or bondtype == 'WARR' or bondtype =='MOB':
                break
            else:
                bondtype = input('STY or PRF or WARR or MOB: ').upper()
    
        self.bond_type = bondtype

    def get_or_no(self):
        self.or_no = input('OR number: ')


def make_bond_prf(bond):
    wb = load_workbook('Template_PRF.xlsx')
    ws = wb['PRF']
    
    first_prf, second_prf = fitintobox(25, bond.coverage_in_words)
    first_line_prf = turntosentence(first_prf)
    second_line_prf = turntosentence(second_prf)
    liss = Bond.con_prop.get(bond.contractor)
    ws['J10'].value = 'G(13)-' + bond.bond_no
    ws['D16'].value = bond.contractor
    ws['H16'].value =  liss[1].upper()
    ws['D21'].value = bond.agency
    ws['I21'].value = first_line_prf.upper()
    ws['A22'].value = second_line_prf.upper() + "PESOS ONLY"
    ws['F22'].value = bond.coverage
    ws['B30'].value = bond.contract_name    
    ws['B37'].value = bond.coverage_in_words.upper() + "PESOS ONLY"
    ws['B38'].value = f"(Php{bond.coverage:,d}.00; Philippine Currencies)"
    ws['C49'].value = f"{bond.date}, {str(year + 1)}"    
    ws['F56'].value = f"{bond.date}, {str(year)}"
    ws['C63'].value = liss[0].upper()
    ws['G63'].value = manager
    ws['G64'].value = f"Prop/Manager-{bond.contractor}"
    


    ws = wb['ACK']

    first_ack, second_ack = fitintobox(28, bond.coverage_in_words)
    first_line_ack = turntosentence(first_ack)
    second_line_ack = turntosentence(second_ack)

    ws['J3'].value = 'G(13)-' + bond.bond_no
    ws['G6'].value = f"{bond.date}, {str(year)}"
    ws['B13'].value = liss[0].upper()
    ws['G13'].value = liss[1].upper()
    ws['C23'].value = f"{bond.date}, {str(year)}"
    ws['A44'].value = first_line_ack.upper()    
    ws['A45'].value = second_line_ack.upper() + "PESOS ONLY"    
    ws['F44'].value = bond.coverage
    ws['H54'].value = f"{bond.date}, {str(year)}"

    ws = wb['IND']


    first_ind, second_ind = fitintobox(35, bond.coverage_in_words)
    first_line_ind = turntosentence(first_ind)
    second_line_ind = turntosentence(second_ind)

    rating_amount = (rating_prf(bond.coverage))
    rating_amount_str = str(rating_amount)
    amount_split = rating_amount_str.split(".")
    rating_in_words = getWords(int(amount_split[0]))
    try:
        if int(amount_split[1][2]) > 4:
            fraction = amount_split[1][0] + str(int(amount_split[1][1]) + 1)
        else:
            fraction = amount_split[1][0:2]
    except:    
        fraction = amount_split[1][0:2]


    ws['J1'].value = 'G(13)-' + bond.bond_no
    ws['B17'].value = bond.contractor
    ws['G17'].value = first_line_ind.upper()
    ws['B18'].value = second_line_ind.upper() + 'PESOS ONLY'
    ws['F18'].value = bond.coverage
    ws['B19'].value = bond.agency.upper()
    ws['G24'].value = rating_in_words.upper()
    ws['B25'].value = f'AND {fraction}/100 ONLY'
    ws['F25'].value = rating_amount

    ws = wb['INDBK']    


    ws['I24'].value = f"{bond.date}, {str(year)}"
    ws['C26'].value = liss[0].upper() #PROPIETOR
    ws['C27'].value = bond.contractor
    ws['H26'].value = liss[1].upper() #ADDRESS
    ws['B52'].value = liss[0].upper() #PROPIETOR
    ws['G52'].value = liss[1].upper() #ADDRESS
    ws['H43'].value = f"{bond.date}, {str(year)}"
    ws = wb['OR']

    payment = getpayment(rating_amount)
    payment_str = str(payment)
    pay_list = payment_str.split(".")
    payment_whole = int(pay_list[0])
    try:
        if int(pay_list[1][2]) > 4:
            fraction_or = pay_list[1][0] + str(int(pay_list[1][1]) + 1)
        else:
            fraction_or = pay_list[1][0:2]
    except:    
        fraction_or = pay_list[1][0:2]

    first_or, second_or = fitintobox(30,getWords(payment_whole))
    first_line_or = turntosentence(first_or)
    second_line_or = turntosentence(second_or) +'PESOS ' f'and {fraction_or}/100 ONLY'


    ws['E17'].value = rating_amount
    ws['A5'].value = f"{bond.date}, {str(year)}"
    ws['B6'].value = bond.contractor
    ws['A8'].value = liss[1]
    ws['B12'].value = first_line_or.upper()
    ws['B13'].value = second_line_or.upper()
    ws['B15'].value = 'G(13)-' + bond.bond_no
    ws['B16'].value = f'{bond.date},{str(year)} - {bond.date},{str(year + 1)}'
    wb.save(f'G(13)-{bond.bond_no}--CIC.{bond.contractor}-PRF.xlsx')


def make_bond_sty(bond):

    wb = load_workbook('Template_STY.xlsx')
    ws = wb['STY']
    
    liss = Bond.con_prop.get(bond.contractor)
    first_sty, second_sty = fitintobox(25, bond.coverage_in_words)
    first_line_sty = turntosentence(first_sty)
    second_line_sty = turntosentence(second_sty)
    bond.prop
    ws['J9'].value = 'G(16)-' + bond.bond_no
    ws['C16'].value = bond.contractor + '-' + bond.prop_address 
    ws['E20'].value = bond.agency
    ws['H21'].value = first_line_sty.upper()
    ws['B22'].value = second_line_sty.upper() + "PESOS ONLY"
    ws['G22'].value = bond.coverage
    ws['B30'].value = bond.contract_name    
    ws['B37'].value = bond.coverage_in_words.upper() + "PESOS ONLY"
    ws['B38'].value = f"(Php{bond.coverage:,d}.00; Philippine Currencies)"
    ws['D49'].value = f"{bond.date}, {str(year + 1)}"
    ws['F56'].value = f"{bond.date}, {str(year)}" 
    ws['C63'].value = bond.prop
    ws['H63'].value = manager

    ws = wb['ACK'] 

    first_ack, second_ack = fitintobox(35, bond.coverage_in_words)
    first_line_ack = turntosentence(first_ack)
    second_line_ack = turntosentence(second_ack)

    ws['J4'].value = 'G(16)-' + bond.bond_no
    ws['G8'].value = f"{bond.date}, {str(year)}" 
    ws['B14'].value = liss[0].upper()
    ws['G14'].value = liss[1].upper()
    ws['C23'].value = f"{bond.date}, {str(year)}" 
    ws['A44'].value = first_line_ack.upper()
    ws['A45'].value = second_line_ack.upper() + "PESOS ONLY"
    ws['F44'].value = bond.coverage
    ws['H57'].value = f"{bond.date}, {str(year)}" 
    
    ws = wb['IND']

    first_ind, second_ind = fitintobox(40, bond.coverage_in_words)
    first_line_ind = turntosentence(first_ind)
    second_line_ind = turntosentence(second_ind)

    rating_amount = (rating_sty(bond.coverage))
    rating_amount_str = str(rating_amount)
    amount_split = rating_amount_str.split(".")
    rating_in_words = getWords(int(amount_split[0]))

    try:
        if int(amount_split[1][2]) > 4:
            fraction = amount_split[1][0] + str(int(amount_split[1][1]) + 1)
        else:
            fraction = amount_split[1][0:2]
    except:    
        fraction = amount_split[1][0:2]

    ws['J1'].value = 'G(16)-' + bond.bond_no
    ws['A17'].value = bond.contractor
    ws['G17'].value = first_line_ind.upper()
    ws['A18'].value = second_line_ind.upper() + 'PESOS ONLY'
    ws['F18'].value = bond.coverage
    ws['A19'].value = bond.agency.upper()
    ws['G24'].value = rating_in_words.upper()
    ws['B25'].value = f'AND {fraction}/100 ONLY'
    ws['F25'].value = rating_amount

    ws = wb['INDBK']

    ws['I24'].value = f"{bond.date}, {str(year)}" 
    ws['B26'].value = liss[0].upper() #PROPIETOR
    ws['B27'].value = bond.contractor
    ws['H26'].value = liss[1].upper() #ADDRESS
    ws['H44'].value = f"{bond.date}, {str(year)}" 
    ws['B52'].value = liss[0].upper() #PROPIETOR
    ws['G52'].value = liss[1].upper() #ADDRESS

    ws = wb['OR']

    payment = getpayment(rating_amount)
    payment_str = str(payment)
    pay_list = payment_str.split(".")
    payment_whole = int(pay_list[0])
    try:
        if int(pay_list[1][2]) > 4:
            fraction_or = pay_list[1][0] + str(int(pay_list[1][1]) + 1)
        else:
            fraction_or = pay_list[1][0:2]
    except:    
        fraction_or = pay_list[1][0:2]

    first_or, second_or = fitintobox(30,getWords(payment_whole))
    first_line_or = turntosentence(first_or)
    second_line_or = turntosentence(second_or) +'PESOS ' f'and {fraction_or}/100 ONLY'

    ws['A5'].value = f"{bond.date}, {str(year)}" 
    ws['B6'].value = bond.contractor
    ws['B8'].value = liss[1]
    ws['B12'].value = first_line_or.upper()
    ws['B13'].value = second_line_or.upper()
    ws['B15'].value = 'G(16)-' + bond.bond_no
    ws['B16'].value = f"{bond.date}, {str(year)}  - {bond.date},{str(year + 1)}"
    ws['E17'].value = rating_amount
    
    wb.save(f'G(16)-{bond.bond_no}--CIC.{bond.contractor}-STY.xlsx')

def make_bond_mob(bond):

    wb = load_workbook('Template_MOB.xlsx')
    ws = wb['MOB']
    
    liss = Bond.con_prop.get(bond.contractor)
    first_mob, second_mob = fitintobox(25, bond.coverage_in_words)
    first_line_mob = turntosentence(first_mob)
    second_line_mob = turntosentence(second_mob)

    ws['J9'].value = 'G(16)-' + bond.bond_no
    ws['C15'].value = bond.contractor + '-' + liss[1].upper() #address of contractor
    ws['E20'].value = bond.agency
    ws['H21'].value = first_line_mob.upper()
    ws['A22'].value = second_line_mob.upper() + "PESOS ONLY"
    ws['G22'].value = bond.coverage    
    ws['B30'].value = bond.contract_name
    ws['B37'].value = bond.coverage_in_words.upper() + "PESOS ONLY"
    ws['B38'].value = f"(Php{bond.coverage:,d}.00; Philippine Currencies)"
    ws['D49'].value = f"{bond.date}, {str(year + 1)}"
    ws['F56'].value = f"{bond.date}, {str(year)}"
    ws['B63'].value = liss[0].upper()

    
    ws = wb['ACK'] 

    first_ack, second_ack = fitintobox(28, bond.coverage_in_words)
    first_line_ack = turntosentence(first_ack)
    second_line_ack = turntosentence(second_ack)

    ws['J4'].value = 'G(16)-' + bond.bond_no
    ws['G8'].value = f"{bond.date}, {str(year)}"
    ws['B14'].value = liss[0].upper()
    ws['H14'].value = liss[1].upper()
    ws['C22'].value = f"{bond.date}, {str(year)}"
    ws['A44'].value = first_line_ack.upper()
    ws['A45'].value = second_line_ack.upper() + "PESOS ONLY"
    ws['F44'].value = bond.coverage
    ws['H57'].value = f"{bond.date}, {str(year)}"
    
    ws = wb['IND']

    first_ind, second_ind = fitintobox(40, bond.coverage_in_words)
    first_line_ind = turntosentence(first_ind)
    second_line_ind = turntosentence(second_ind)

    rating_amount = (rating_sty(bond.coverage))
    rating_amount_str = str(rating_amount)
    amount_split = rating_amount_str.split(".")
    rating_in_words = getWords(int(amount_split[0]))

    try:
        if int(amount_split[1][2]) > 4:
            fraction = amount_split[1][0] + str(int(amount_split[1][1]) + 1)
        else:
            fraction = amount_split[1][0:2]
    except:    
        fraction = amount_split[1][0:2]

    ws['J1'].value = 'G(16)-' + bond.bond_no
    ws['B17'].value = bond.contractor
    ws['G17'].value = first_line_ind.upper()
    ws['B18'].value = second_line_ind.upper() + 'PESOS ONLY'
    ws['F18'].value = bond.coverage
    ws['B19'].value = bond.agency.upper()
    ws['G24'].value = rating_in_words.upper()
    ws['B25'].value = f'AND {fraction}/100 ONLY'
    ws['F25'].value = rating_amount

    ws = wb['INDBK']

    ws['I24'].value = f"{bond.date}, {str(year)}"
    ws['C26'].value = liss[0].upper() #PROPIETOR
    ws['C27'].value = bond.contractor
    ws['H26'].value = liss[1].upper() #ADDRESS
    ws['H44'].value = f"{bond.date}, {str(year)}"
    ws['B52'].value = liss[0].upper() #PROPIETOR
    ws['G52'].value = liss[1].upper() #ADDRESS

    ws = wb['OR']

    payment = getpayment(rating_amount)
    payment_str = str(payment)
    pay_list = payment_str.split(".")
    payment_whole = int(pay_list[0])
    try:
        if int(pay_list[1][2]) > 4:
            fraction_or = pay_list[1][0] + str(int(pay_list[1][1]) + 1)
        else:
            fraction_or = pay_list[1][0:2]
    except:    
        fraction_or = pay_list[1][0:2]

    first_or, second_or = fitintobox(30,getWords(payment_whole))
    first_line_or = turntosentence(first_or)
    second_line_or = turntosentence(second_or) +'PESOS ' f'and {fraction_or}/100 ONLY'

    ws['A5'].value = f"{bond.date}, {str(year)}"
    ws['B6'].value = bond.contractor
    ws['B8'].value = liss[1]
    ws['B12'].value = first_line_or.upper()
    ws['B13'].value = second_line_or.upper()
    ws['B15'].value = 'G(16)-' + bond.bond_no
    ws['B16'].value = f'{bond.date},{str(year)} - {bond.date},{str(year + 1)}'
    ws['E17'].value = rating_amount

    wb.save(f'G(16)-{bond.bond_no}--CIC.{bond.contractor}-MOB.xlsx')

def make_bond_warr(bond):

    wb = load_workbook('Template_WARR.xlsx')
    ws = wb['WARR']
    
    liss = Bond.con_prop.get(bond.contractor)
    first_warr, second_warr = fitintobox(25, bond.coverage_in_words)
    first_line_warr = turntosentence(first_warr)
    second_line_warr = turntosentence(second_warr)

    ws['J10'].value = 'G(16)-' + bond.bond_no
    ws['D16'].value = bond.contractor + '-' + liss[1].upper() #address of contractor
    ws['E21'].value = bond.agency
    ws['H22'].value = first_line_warr.upper()
    ws['A23'].value = second_line_warr.upper() + "PESOS ONLY"
    ws['G23'].value = bond.coverage    
    ws['D33'].value = bond.contract_name
    ws['B38'].value = bond.coverage_in_words.upper() + "PESOS ONLY"
    ws['B39'].value = f"(Php{bond.coverage:,d}.00; Philippine Currencies)"
    ws['D50'].value = f"{bond.date}, {str(year + 1)}"
    ws['F57'].value = f"{bond.date}, {str(year)}"
    ws['B64'].value = liss[0].upper()

    
    ws = wb['ACK'] 

    first_ack, second_ack = fitintobox(28, bond.coverage_in_words)
    first_line_ack = turntosentence(first_ack)
    second_line_ack = turntosentence(second_ack)

    ws['J3'].value = 'G(16)-' + bond.bond_no
    ws['G7'].value = f"{bond.date}, {str(year)}"
    ws['B13'].value = liss[0].upper()
    ws['H13'].value = liss[1].upper()
    ws['C22'].value = f"{bond.date}, {str(year)}"
    ws['A44'].value = first_line_ack.upper()
    ws['A45'].value = second_line_ack.upper() + "PESOS ONLY"
    ws['F44'].value = bond.coverage
    ws['H57'].value = f"{bond.date}, {str(year)}"
    
    ws = wb['IND']

    first_ind, second_ind = fitintobox(40, bond.coverage_in_words)
    first_line_ind = turntosentence(first_ind)
    second_line_ind = turntosentence(second_ind)

    rating_amount = (rating_sty(bond.coverage))
    rating_amount_str = str(rating_amount)
    amount_split = rating_amount_str.split(".")
    rating_in_words = getWords(int(amount_split[0]))

    try:
        if int(amount_split[1][2]) > 4:
            fraction = amount_split[1][0] + str(int(amount_split[1][1]) + 1)
        else:
            fraction = amount_split[1][0:2]
    except:    
        fraction = amount_split[1][0:2]

    ws['J1'].value = 'G(16)-' + bond.bond_no
    ws['B17'].value = bond.contractor
    ws['G17'].value = first_line_ind.upper()
    ws['B18'].value = second_line_ind.upper() + 'PESOS ONLY'
    ws['F18'].value = bond.coverage
    ws['B19'].value = bond.agency.upper()
    ws['G24'].value = rating_in_words.upper()
    ws['B25'].value = f'AND {fraction}/100 ONLY'
    ws['F25'].value = rating_amount

    ws = wb['INDBK']

    ws['I24'].value = f"{bond.date}, {str(year)}"
    ws['C26'].value = liss[0].upper() #PROPIETOR
    ws['C27'].value = bond.contractor
    ws['H26'].value = liss[1].upper() #ADDRESS
    ws['H44'].value =f"{bond.date}, {str(year)}"
    ws['B52'].value = liss[0].upper() #PROPIETOR
    ws['G52'].value = liss[1].upper() #ADDRESS

    ws = wb['OR']

    payment = getpayment(rating_amount)
    payment_str = str(payment)
    pay_list = payment_str.split(".")
    payment_whole = int(pay_list[0])
    try:
        if int(pay_list[1][2]) > 4:
            fraction_or = pay_list[1][0] + str(int(pay_list[1][1]) + 1)
        else:
            fraction_or = pay_list[1][0:2]
    except:    
        fraction_or = pay_list[1][0:2]

    first_or, second_or = fitintobox(30,getWords(payment_whole))
    first_line_or = turntosentence(first_or)
    second_line_or = turntosentence(second_or) +'PESOS ' f'and {fraction_or}/100 ONLY'

    ws['A5'].value = f"{bond.date}, {str(year)}"
    ws['B6'].value = bond.contractor
    ws['B10'].value = liss[1]
    ws['B12'].value = first_line_or.upper()
    ws['B13'].value = second_line_or.upper()
    ws['B17'].value = 'G(16)-' + bond.bond_no
    ws['B18'].value = f'{bond.date},{str(year)} - {bond.date},{str(year + 1)}'
    ws['F18'].value = rating_amount

    wb.save(f'G(16)-{bond.bond_no}--CIC.{bond.contractor}-WARR.xlsx')







def report_prf(bond):
    wb = load_workbook('CIC_MONTHLY_REPORT_PRF.xlsx')
    ws = wb['REPORT']
    cells = []
    for row in ws.iter_rows(min_row = 2, max_col = 7, max_row = 100):
        for cell in row:
            if cell.value is not None:
                continue
            else:
                cells.append(cell)
                
                   
    cells = cells[0:7]
    cells[0].value = f"{bond.date}, {str(year)}"
    if bond.bond_type == 'PRF':
        cells[1].value = f'G(13){bond.bond_no}'
    else:
        cells[1].value = f'G(16){bond.bond_no}'
    cells[2].value = bond.or_no
    cells[3].value = bond.contract_name
    cells[4].value = bond.contractor
    if bond.bond_type == 'PRF':
        cells[5].value = rating_prf(bond.coverage)
    else:
        cells[5].value = rating_sty(bond.coverage)
    if bond.bond_type == 'PRF':
        cells[6].value = getpayment(rating_prf(bond.coverage))
    else:
        cells[6].value = getpayment(rating_sty(bond.coverage))
    
    wb.save('CIC_MONTHLY_REPORT_PRF.xlsx')


def remittance_prf(bond):
    wb = load_workbook('REMITTANCE_PRF.xlsx')
    ws = wb['REMIT']
    cells = []
    for row in ws.iter_rows(min_row = 2, min_col = 2, max_col = 5 , max_row = 100):
        for cell in row:
            if cell.value is not None:
                continue
            else:
                cells.append(cell)

    cells = cells[0:4]
    
    cells[0].value = f"{bond.date}, {str(year)}"
    if bond.bond_type == 'PRF':
        cells[1].value = f'G(13){bond.bond_no}'
    else:
        cells[1].value = f'G(16){bond.bond_no}'
    
    cells[2].value = bond.or_no
    if bond.bond_type == 'PRF':
        cells[3].value = rating_prf(bond.coverage)
    else:
        cells[3].value = rating_sty(bond.coverage)

    wb.save('REMITTANCE_PRF.xlsx')

def report_sty(bond):
    wb = load_workbook('CIC_MONTHLY_REPORT_STY.xlsx')
    ws = wb['REPORT']
    cells = []
    for row in ws.iter_rows(min_row = 2, max_col = 7, max_row = 100):
        for cell in row:
            if cell.value is not None:
                continue
            else:
                cells.append(cell)
                
                   
    cells = cells[0:7]
    cells[0].value = f"{bond.date}, {str(year)}"
    if bond.bond_type == 'PRF':
        cells[1].value = f'G(13){bond.bond_no}'
    else:
        cells[1].value = f'G(16){bond.bond_no}'
    cells[2].value = bond.or_no
    cells[3].value = bond.contract_name
    cells[4].value = bond.contractor
    if bond.bond_type == 'PRF':
        cells[5].value = rating_prf(bond.coverage)
    else:
        cells[5].value = rating_sty(bond.coverage)
    if bond.bond_type == 'PRF':
        cells[6].value = getpayment(rating_prf(bond.coverage))
    else:
        cells[6].value = getpayment(rating_sty(bond.coverage))
    
    wb.save('CIC_MONTHLY_REPORT_STY.xlsx')


def remittance_sty(bond):
    wb = load_workbook('REMITTANCE_STY.xlsx')
    ws = wb['REMIT']
    cells = []
    for row in ws.iter_rows(min_row = 2, min_col = 2, max_col = 5 , max_row = 100):
        for cell in row:
            if cell.value is not None:
                continue
            else:
                cells.append(cell)

    cells = cells[0:4]
    
    cells[0].value = f"{bond.date}, {str(year)}"
    if bond.bond_type == 'PRF':
        cells[1].value = f'G(13){bond.bond_no}'
    else:
        cells[1].value = f'G(16){bond.bond_no}'
    
    cells[2].value = bond.or_no
    if bond.bond_type == 'PRF':
        cells[3].value = rating_prf(bond.coverage)
    else:
        cells[3].value = rating_sty(bond.coverage)

    wb.save('REMITTANCE_STY.xlsx')
    
def what_to_make(number):
    bonds = [Bond() for i in range(1, number + 1)]
    for bnd in bonds:  
        if bnd.bond_type == 'STY':
            make_bond_sty(bnd)
            report_sty(bnd)
            remittance_sty(bnd)
        elif bnd.bond_type == 'PRF':
            make_bond_prf(bnd)
            report_prf(bnd)
            remittance_prf(bnd)
        elif bnd.bond_type == 'WARR':
            make_bond_warr(bnd)
            report_sty(bnd)
            remittance_sty(bnd)
        elif bnd.bond_type == 'MOB':
            make_bond_mob(bnd)
            report_sty(bnd)
            remittance_sty(bnd)
            
    

def make_bond():
    
    while True:
        number = input(f'How many bonds are to be made? ')
        try:
            number = int(number)
        except:
            print('Invalid, try again.')

        if type(number) is int:
            break
        else:
            number = input(f'How many bonds are to be made? ')
            if type(number) is int:
                pass
            else:
                print('Invalid, try again.')
          
    return number    


def edit_config():
    config_path = 'config.json'
    # Load current config
    with open(config_path, 'r') as f:
        config = json.load(f)
    print("Current config values:")
    for key, value in config.items():
        print(f"{key}: {value}")
    print("\nEnter new values (leave blank to keep current):")
    for key in config:
        new_value = input(f"{key} [{config[key]}]: ")
        if new_value.strip() != "":
            config[key] = new_value
    # Save updated config
    with open(config_path, 'w') as f:
        json.dump(config, f, indent=4)
    print("Config updated.")


def main():
    print("1. Make bond. \n" 
    "2. Update CTC")
    choice = int(input(f"What do you want to do? "))
    if choice == 1:
        what_to_make(make_bond())
    elif choice == 2:
        edit_config()

main()